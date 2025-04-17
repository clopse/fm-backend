import os
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font

def update_energy_excel(filepath, data):
    headers = [
        "Billing Start", "Billing End", "Invoice Date", "Invoice #", "GPRN", "MIC",
        "Day kWh", "Night kWh", "Total kWh", "Subtotal €", "Total €",
        "Day Rate", "Night Rate", "Avg Rate €/kWh", "File Path", "Supplier"
    ]

    if not os.path.exists(filepath):
        wb = Workbook()
        ws = wb.active
        ws.title = "Energy Bills"
        ws.append(headers)
        for col in range(1, len(headers) + 1):
            ws[get_column_letter(col) + "1"].font = Font(bold=True)
        wb.save(filepath)

    wb = load_workbook(filepath)
    ws = wb.active

    # Check if entry already exists based on billing period
    for row in ws.iter_rows(min_row=2, values_only=True):
        if str(row[0]) == data["Billing Start"] and str(row[1]) == data["Billing End"]:
            print("⚠️ Entry already exists in Excel. Skipping.")
            return

    ws.append([
        data.get("Billing Start", ""),
        data.get("Billing End", ""),
        data.get("Invoice Date", ""),
        data.get("Invoice #", ""),
        data.get("GPRN", ""),
        data.get("MIC", ""),
        data.get("Day kWh", ""),
        data.get("Night kWh", ""),
        data.get("Total kWh", ""),
        data.get("Subtotal €", ""),
        data.get("Total €", ""),
        data.get("Day Rate", ""),
        data.get("Night Rate", ""),
        data.get("Avg Rate €/kWh", ""),
        data.get("File Path", ""),
        data.get("Supplier", "")
    ])

    wb.save(filepath)
    print(f"📝 Excel updated: {filepath}")
